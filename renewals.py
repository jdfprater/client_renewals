#! python3

import openpyxl, time, pprint
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

start_time = time.time()
print('Starting...')

def isInteger(value):
    try:
        int(value)
        return True
    except ValueError:
        return False

GYMS =              ['TX-AUSTIN ANDERSON ARBOR',
                      'TX-AUSTIN CEDAR PARK',
                      'TX-AUSTIN CYPRESS CREEK',
                      'TX-AUSTIN HESTERS CROSSING',
                      'TX-AUSTIN NORTH ROUND ROCK',
                      'TX-AUSTIN TECHRIDGE',
                      'TX-GEORGETOWN',
                      'TX-PFLUGERVILLE']


renewals = {}

print('Calculating appointments...')
appointments = openpyxl.load_workbook('Member Appointments.xlsx')
ap_sheet = appointments['Sheet1']

ID_col = column_index_from_string('E')
service_provider = column_index_from_string('D')
service_type = column_index_from_string('B')
service = column_index_from_string('C')

for row in ap_sheet.rows:
    ID = str(row[ID_col-1].value)
    trainer = row[service_provider-1].value
    personal_training = row[service_type-1].value
    session = row[service-1].value
    if personal_training:
        if 'Personal' in personal_training and 'PT' in session:
            if ID in renewals:
                renewals[ID]['Appointments'] += 1
            else:
                renewals.update( {ID : {'Gym Name' : '', 'Member Name' : '', 'PT' : trainer, 'Sessions Remaining' : 0, 'Appointments' : 1, 'Has EFT' : ''} } )
    else:
        continue


active_clients = openpyxl.load_workbook('PT Business Report - Active PT 1-on-1 Detail.xlsx')
ac_sheet = active_clients['Active PT 1-on-1 Detail Report']

gym_name = column_index_from_string('D')
member_ID = column_index_from_string('F')
client_name = column_index_from_string('H')
pt_name = column_index_from_string('K')
is_pif = column_index_from_string('N')
has_sessions = column_index_from_string('S')
pif_type = ['Session', 'GF', 'PIF']
cancel_request = column_index_from_string('AB')


print('Calculating Sessions...')
for row in ac_sheet.rows:
    gym = row[gym_name-1].value
    ID = str(row[member_ID-1].value)
    name = row[client_name-1].value
    pt = row[pt_name-1].value
    pif = row[is_pif-1].value
    sessions = row[has_sessions-1].value
    if ID:
        if ID in renewals and any(x in (gym) for x in GYMS): #and any(x in (pif) for x in pif_type)
            renewals[ID]['Gym Name'] = gym
            renewals[ID]['Member Name'] = name
            renewals[ID]['Sessions Remaining'] += sessions
        elif ID not in renewals and any(x in (gym) for x in GYMS): #and any(x in (pif) for x in pif_type)
            renewals.update( {ID : {'Gym Name' : gym, 'Member Name' : name, 'PT' : pt, 'Sessions Remaining' : sessions, 'Appointments' : 0, 'Has EFT' : ''} } )
    else:
        continue

print('Removing instances without names...')
renewals = { k : v for k,v in renewals.items() if v['Member Name'] }


print('Searching for clients with EFT agreements...')
for row in ac_sheet.rows:
    ID = str(row[member_ID-1].value)
    eft = row[is_pif-1].value
    cancel = row[cancel_request-1].value
    if 'EFT' in eft and ID in renewals and not cancel:
        renewals[ID]['Has EFT'] = 'Y'


print('Writing...')
headers = ['ID', 'Gym', 'Name', 'PT', 'Sessions', 'Appointments', 'Has EFT']

wb = Workbook()
sheet = wb.active

for i in range(len(headers)):
    sheet.cell(row=1, column=i+1).font = Font(bold=True)
    sheet.cell(row=1, column=i+1).value = headers[i]

row = 2
for ID, data in renewals.items():
    sheet.cell(row=row, column=1, value=ID)
    column = 2
    for k,v in data.items():
        sheet.cell(row=row, column=column, value=v)
        column +=1
    row += 1

for i in range(len(renewals)):
    if int(sheet.cell(row=i+2, column=5).value) < 5:
        sheet.cell(row=i+2, column=5).fill = PatternFill(fgColor='FFC7CE', fill_type = 'solid')


wb.save('renewals.xlsx')
wb.close()

print("--- %s seconds ---" % (time.time() - start_time))
